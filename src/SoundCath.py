
from tkinter import StringVar, ttk, Tk, IntVar, Toplevel
import time
import TesterBackend as tb
from concurrent import futures
import csv
import os
import threading
import openpyxl
from openpyxl.styles import fills, colors
from copy import copy

pulseecho_trigger_delay = 5.0 # The delay between the starting and capture of the waveform
channel_delay = 1.0 # time between channels in all channels

class TesterFrontEnd:  # a GUI front end for the test

    def __init__(self):
        # Configure the GUI Basic Parts
        self.root = Tk() # setup the GUI base
        self.root.title("Catheter Tester Script")
        self.root.geometry("600x400")

        self.style = ttk.Style()
        self.style.configure("TCheckbutton", font = ("Arial", 12))
        self.style.configure("TButton", font = ("Arial", 12))
        self.style.configure("TEntry", font = ("Arial", 24))
        self.style.configure("TLabel", font = ("Arial", 12))
        self.style.configure("Green.TLabel", foreground="green", font = ("Arial", 14))

        # All of the Internal Variables
        self.channel = -1 # the channel we are in
        self.promptcapture = IntVar(self.root, 1)
        self.impedancetest = IntVar(self.root, 0) # bool for if we are gonna run the impedance test
        self.allchannels = IntVar(self.root, 0) # bool if we are gonna test all channels
        self.pulseechotest = IntVar(self.root, 0) # bool if we are gonna run the pulse echo test
        self.dongletest = IntVar(self.root, 0) # bool if we are gonna run a dongle test
        self.text = StringVar(self.root, "Channel " + str(self.channel + 1)) # the string to display the channel

        # Variable to Store the Results of all of the Tests
        self.passmap = None
        self.backend = tb.CatheterTester() # Backend tester that does the actual work 
        self.triggered = IntVar(self.root, 0)
        self.stopped = IntVar(self.root, 0)

        self.window = ttk.Frame(self.root)  # All widget elements
        self.pool = futures.ThreadPoolExecutor(max_workers=1)
        
        self.upbutton = ttk.Button(self.root, text = "Up", command = self.IncChannel) # the up button
        self.downbutton = ttk.Button(self.root, text = "Down", command = self.DecChannel) # the down button
        self.stopbutton = ttk.Button(self.root, text = "Stop", command = lambda: self.stopped.set(1))
        self.capturebutton = ttk.Button(self.root, text = "Capture", command = lambda: self.triggered.set(1))
        
        self.label = ttk.Label(self.root, textvariable = self.text, width = 11) # Label to draw the Channel Number
        self.filenamelabel = ttk.Label(self.root, text = "File Name to Save:") # Entry box to put the filename into
        self.readylabel = ttk.Label(self.root, text = "Ready", style="Green.TLabel")
        self.filename = ttk.Entry(self.root)

        self.impedancetestbutton = ttk.Checkbutton(self.root, variable = self.impedancetest, text = "Impedance Test") # all of the check boxes
        self.allchannelsbutton  = ttk.Checkbutton(self.root, variable = self.allchannels,  text = "Run All Channels")
        self.pulseechotestbutton = ttk.Checkbutton(self.root, variable = self.pulseechotest, text = "Pulse Echo")
        self.dongletestbutton = ttk.Checkbutton(self.root, variable = self.dongletest, text = "Dongle Test")
        self.promptcapturebutton = ttk.Checkbutton(self.root, variable = self.promptcapture, text = "Prompt Capture")

        self.channels = IntVar(self.root, 64)
        self.channelselectlabel = ttk.Label(self.root, text="Channels")
        self.channel32select = ttk.Radiobutton(self.root, variable = self.channels, text="32", value=32)
        self.channel64select = ttk.Radiobutton(self.root, variable = self.channels, text="64", value=64)
        self.channel96select = ttk.Radiobutton(self.root, variable = self.channels, text="96", value=96)

        self.runbutton = ttk.Button(self.root, text = "Run Tests", command = self.run_button_cb) # all of the buttons to actually run the tests and get results and stuff
        self.reportbutton = ttk.Button(self.root, text = "Generate Report", command = self.GenerateXLSXReport)

    def Draw(self) -> None: # positions and draws all of the widgets in the frame

        self.upbutton.place(x = 400, y = 50)
        self.label.place(x = 240 , y = 50)
        self.downbutton.place(x = 50, y = 50)
        
        self.impedancetestbutton.place(x = 50, y = 100)
        self.pulseechotestbutton.place(x = 250, y = 100)
        self.allchannelsbutton.place(x = 400, y = 100)
        self.dongletestbutton.place(x = 50, y = 150)
        self.promptcapturebutton.place(x = 250, y = 150)
        self.readylabel.place(x = 400, y = 150)

        self.channelselectlabel.place(x = 50, y = 200)
        self.channel32select.place(x = 50, y = 230)
        self.channel64select.place(x = 250, y = 230)
        self.channel96select.place(x = 400, y = 230)

        self.reportbutton.place(x = 50, y = 340)
        self.runbutton.place(x = 400, y = 340)
        self.filenamelabel.place(x = 50, y = 280)
        self.filename.place(x = 300, y = 280, height = 50, width = 250)
        self.window.mainloop()
    
    def run_button_cb(self):
        self.pool.submit(self.RunTests)
    
    def RunSingleChannelTest(self, channel, filename) -> None:

        print(f"Running Test on Channel: {channel + 1}")
    
        if self.pulseechotest.get() != 0:
            self.passmap["PulseEcho"][channel - 1] = self.RunPulseEchoTest(channel, filename) # Run the Tests and record the Results
        
        if self.impedancetest.get() != 0:
            self.passmap["Impedance"][channel - 1] = self.RunImpedanceTest(channel, filename) 

        if self.dongletest.get() != 0:
            self.passmap["Dongle"][channel - 1] =  self.RunDongleTest(channel, filename) 
    
    def RunTests(self) -> None:
        
        print("Running Tests")
        if(self.impedancetest.get() == 0 and self.dongletest.get() == 0 and self.pulseechotest.get() == 0): # repeat the same process with the impedance test/dongle test
            return
        
        self.root.after(0, lambda: self.readylabel.place_forget())
        
        mc = int(self.channels.get())
        self.passmap = { "Impedance": [[False, None, None]] * mc, "PulseEcho": [[False, None, None, None]] * mc, "Dongle": [[False, None, None]] * mc }
    
        filename = os.getcwd() + '\\' + self.filename.get()
        path = "\\".join(filename.split("\\")[0:-1]) # create the path of the file if it wasn't already there

        os.makedirs(path, exist_ok=True) # if the directory isn't there create it

        if self.promptcapture.get() and self.pulseechotest.get():   # if we want to prompt capture then open up a capture window
            self.root.after(0, lambda: self.capturebutton.place(x =180, y = 340))
        
        if self.allchannels.get() != 0:
            
            print("Testing All Channels")
            
            self.root.after(0, lambda: self.stopbutton.place(x = 290, y = 340))
   
            for i in range(0, mc):
                if self.stopped.get():
                    break
                self.RunSingleChannelTest(i, filename)
                time.sleep(channel_delay)   

            self.root.after(0, lambda: self.stopbutton.place_forget())
            self.stopped.set(0)
            
            try: 
                self.GenerateXLSXReport() # after all channels have been run then generate a report of the findings
            except Exception as e:
                print(e)
                
        else:
            self.RunSingleChannelTest(self.channel, filename) # run only one channel

        if self.promptcapture.get() and self.pulseechotest.get(): # after we are done destroy the capture window if we are using it
            self.root.after(0, lambda: self.capturebutton.place_forget())
        
        self.root.after(0, lambda: self.readylabel.place(x = 400, y = 150))
        print("Done Running Tests")
            
    def CapturePopup(self) -> Toplevel: 
        
        window = Toplevel() # a secondary window
        
        def CaptureButtonCB() -> None:
            self.triggered.set(1) # mark the channel is triggered
                
        button = ttk.Button(window, text = "Capture", command = CaptureButtonCB) # create a button, when pushed trigger the scope to capture
        button.pack() # place the button in the window
        button.wait_variable(self.triggered) # wait for the button to be pressed
        self.triggered.set(0)

        return window

    def RunImpedanceTest(self, channel, filename) -> tuple[bool, float, float, float]:
        print("Running Impedance Test")
        return self.backend.ImpedanceTest(channel, int(self.channels.get()), filename) # run the test with the filename

    def RunPulseEchoTest(self, channel, filename) -> tuple[bool, float]:    
        if not self.promptcapture.get():
            time.sleep(pulseecho_trigger_delay)
        else:
            self.triggered.set(0) # mark the the channel as untriggered
            while not self.triggered.get(): # wait for the button to be pressed
                pass
            
        print("Running Pulse Echo Test")
        return self.backend.PulseEchoTest(1, channel, int(self.channels.get()), filename) # run the test on scope channel 1 with file name filename
        
    def RunDongleTest(self, channel, filename) -> tuple[bool, float]:
        print("Running Dongle Test")
        return self.backend.DongleTest(channel, int(self.channels.get()), filename) # Run the Dongle test with the Filename as its file name

    def IncChannel(self) -> None: # increments the channel and displays the change
        mc = self.channels.get()
        if(self.channel < mc):
            self.channel += 1
            self.text.set("Channel " + str(self.channel + 1))
            self.backend.SetChannel(self.channel, mc)
            print(self.backend.arduino.ReadLine())

    def DecChannel(self) -> None: # Decrements the channel and display the change
        mc = self.channels.get()
        if(self.channel >= 0):
            self.channel -= 1
            self.text.set("Channel " + str(self.channel + 1))
            self.backend.SetChannel(self.channel, mc)
            print(self.backend.arduino.ReadLine())
            
    def GenerateCSVReport(self) -> None: # Generates a CSV Report with all of the results

        print("Generating CSV Report")
        filename = os.getcwd() + '\\' + self.filename.get() 
        channels = [i + 1 for i in range(tb.max_channel)]

        if self.pulseechotest.get():
            with open(filename + 'PEReport.csv', 'w') as pefile:
                pewriter = csv.writer(pefile)
                
                pewriter.writerow(["Channel", "Passed", "Vpp", "Bandwidth", "Center Frequency"])
                for channel in channels:
                    print(self.passmap["PulseEcho"][channel - 1])
                    data = self.passmap["PulseEcho"][channel - 1]
                    data.insert(0, channel)
                    pewriter.writerow(data)

        if self.impedancetest.get():
            with open(filename + 'ZReport.csv', 'w') as zfile:
                zwriter = csv.writer(zfile)
                
                zwriter.writerow(["Channel", "Passed", "Capacitance"])
                for channel in channels:
                    print(self.passmap["Impedance"][channel - 1])
                    data = self.passmap["Impedance"][channel - 1]
                    data.insert(0, channel)
                    zwriter.writerow(data)
        
        if self.dongletest.get():
            with open(filename + "DongleReport", 'w') as donglefile:
                donglewriter = csv.writer(donglefile)
                
                donglewriter.writerow(["Channel", "Passed", "Capacitance"])
                for channel in channels:
                    print(self.passmap["Dongle"][channel - 1])
                    data = self.passmap["Dongle"][channel - 1]
                    data.insert(0, channel)
                    donglewriter.writerow(data)

    def GenerateXLSXReport(self) -> None:

        print("Generating Excel Report")
        filename = os.getcwd() + '\\' + self.filename.get() + "Report.xlsx" # save the report
        
        append = os.path.isfile(filename) # we are adding overwriting an existing report

        report = None
        if append:
            report = openpyxl.load_workbook(filename)
        else:
            report = openpyxl.Workbook() # create a new excel sheet with multiple pages
        
        # load all of the templates into sheets
        templatepath = os.path.dirname(os.path.abspath(__file__)) + "\\..\\docs\\"
        donglereporttemplate = openpyxl.load_workbook(filename = templatepath + f"DongleTemplate{self.channels.get()}.xlsx")
        impedancereporttemplate = openpyxl.load_workbook(filename = templatepath + f"ImpedanceTemplate{self.channels.get()}.xlsx")
        pereporttemplate = openpyxl.load_workbook(filename = templatepath + f"PETemplate{self.channels.get()}.xlsx")
        
        # copy all of the templates into the new excel sheet and name the sheets accordingly
        red = colors.Color("00FF0000")
        green = colors.Color("0000FF00")
        failcolor = fills.PatternFill(patternType = 'solid', fgColor = red)
        passcolor = fills.PatternFill(patternType = 'solid', fgColor = green)

        mc = self.channels.get()
        
        # fill out the sheet with the data from the pulse echo test results
        if self.pulseechotest.get():        
            pereport = None
            if "Pulse Echo" in report.sheetnames:
                pereport = report["Pulse Echo"]
            else:
                pereport = report.create_sheet("Pulse Echo")
                copy_sheet(pereporttemplate.active, pereport)

            ave_vpp = 0
            ave_bandwidth = 0
            ave_center = 0
            num_samples = 0

            for i in range(8, 8 + mc):
                
                data = self.passmap["PulseEcho"][i - 8] # get the channel test results for PE
                if None in data:
                    continue
                
                pereport["D" + str(i)] = f"{data[1] * 1e3: .2f}" # put the vpp in mV
                pereport["E" + str(i)] = f"{data[3]/data[2] * 100: .2f}" # put the bandwidth in MHz
                pereport["F" + str(i)] = "Pass" if data[0] else "Fail" # if the channel passed put "Pass"
                pereport["G" + str(i)] = "True" if data[1] == 0 else "" # if the channel is dead say so
                pereport["H" + str(i)] = f"{data[2] * 1e-6:.2f}" # put the center frequency in MHz
                
                if data[0]:
                    ave_vpp += data[1]
                    ave_bandwidth += data[3]/data[2] * 100
                    ave_center += data[2] * 1e-6
                    num_samples += 1
                
                color = passcolor if data[0] == True else failcolor # fill the rows with the correct color based on the pass or fail
                rowcells = pereport.iter_cols(min_col = 3, max_col = 8, min_row = i, max_row = i)
                for row in rowcells:
                    for cols in row: 
                        cols.fill = color
            
                pereport["D" + str(9 + mc)] = f"{ave_vpp * 1e3 / num_samples: .2f}" # put the vpp in mV
                pereport["E" + str(9 + mc)] = f"{ave_bandwidth / num_samples: .2f}" # put the bandwidth in MHz
                pereport["H" + str(9 + mc)] = f"{ave_center * 1e-6 / num_samples:.2f}" # put the center frequency in MHz
            
        if self.impedancetest.get():
            impedancereport = None
            if "Impedance" in report.sheetnames:
                impedancereport = report["Impedance"]
            else:
                impedancereport = report.create_sheet("Impedance")
                copy_sheet(impedancereporttemplate.active, impedancereport)
                
            ave_cap = 0
            ave_z = 0
            num_samples = 0
            
            # fill out the sheet with the impedance test results 
            for i in range(11, 11 + mc):    

                data = self.passmap["Impedance"][i - 11]
                if None in data:
                    continue
                
                impedancereport["D" + str(i)] = f"{data[1] * 1e12: .2f}" # put the capacitance in pF
                impedancereport["E" + str(i)] = f"{data[2]}" # put the impedance in ohms
                impedancereport["E" + str(i)] = "Pass" if data[0] else "Fail"  # put if the channel passed or failed
                impedancereport["F" + str(i)] = "Open" if data[1] > 10e-12 and data[1] < 600e-12 else "Short" if data[1] < 0 else "" # Put if the channel is Open or Short based on criteria       
                
                if data[0]:
                    ave_cap += data[1]
                    ave_z += data[2]
                    num_samples += 1
                    
                color = passcolor if data[0] == True else failcolor # highlight the rows with the correct color based on if they passed or failed
                rowcells = impedancereport.iter_cols(min_col = 3, max_col = 7, min_row = i, max_row = i)
                for row in rowcells:
                    for cols in row: 
                        cols.fill = color
                        
                pereport["D" + str(12 + mc)] = f"{ave_cap * 1e12 / num_samples: .2f}" # put the vpp in mV
                pereport["E" + str(12 + mc)] = f"{ave_z / num_samples: .2f}" # put the bandwidth in MHz

        if self.dongletest.get():
            donglereport = None
            if "Dongle" in report.sheetnames:
                dongleresport = report["Dongle"]
            else:
                donglereport = report.create_sheet("Dongle")
                copy_sheet(donglereporttemplate.active, donglereport)
                
            ave_cap = 0
            ave_z = 0
            num_samples = 0
            # fill out the sheet with the dongle test results
            for i in range(11, 11 + mc):    
                
                data = self.passmap["Dongle"][i - 11]
                if None in data:
                    continue
            
                donglereport["D" + str(i)] = f"{data[1] * 1e12: .2f}" # put the capacitance in pF
                donglereport["E" + str(i)] = f"{data[2]}" # impedance in ohms
                donglereport["F" + str(i)] = "Pass" if data[0] else "Fail" # mark if the channel passed or failed
                donglereport["G" + str(i)] = "Open" if data[1] > 10e-12 and data[1] < 180e-12 else "Short" if data[1] < 0 else "" # mark if the channel is open or short based on criteria
                
                if data[0]:
                    ave_cap += data[1]
                    ave_z += data[2]
                    num_samples += 1
                    
                color = passcolor if data[0] == True else failcolor # color the rows according to passing or failing
                rowcells = donglereport.iter_cols(min_col = 3, max_col = 7, min_row = i, max_row = i)
                for row in rowcells:
                    for cols in row: 
                        cols.fill = color
                        
                pereport["D" + str(12 + mc)] = f"{ave_cap * 1e12 / num_samples: .2f}" # put the vpp in mV
                pereport["E" + str(12 + mc)] = f"{ave_z / num_samples: .2f}" # put the bandwidth in MHz
                
        report.save(filename)
        print(f"Saved Report as {filename}")


# Straight from SO on copying whole sheets
def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)

def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)

def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)
            

# the default script starts here

if __name__ == "__main__":

    import atexit
    atexit.register(input, "Press Any Key To Continue")

    gui = TesterFrontEnd()
    gui.Draw()

